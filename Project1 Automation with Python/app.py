

# PROJECT 1: EXCEL AUTOMATION WITH PYTHON

# This script automates the process of updating spreadsheets
# It reads a workbook, calculates a new discounted price for each item,
# writes the new prices into a new column, and generates a Bar Chart!

# Note: You must install the 'openpyxl' package first!
# Terminal: pip install openpyxl

import openpyxl as xl # now we don't need to type openpyxl everytime 
# xl works as alias for it 

from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    # 1. Load the workbook and select the sheet
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Ways to access a particular cell
    # cell = sheet['a1'] 
    # cell = sheet.cell(1, 1)

    # 2. Iterate over all the rows
    # We start at row 2 to skip the header row
    for row in range(2, sheet.max_row + 1):
        
        # The original price is in the 3rd column
        cell = sheet.cell(row, 3)
        
        # Calculate the new price (applying a 10% discount)
        corrected_price = cell.value * 0.9
        
        # 3. Write the new price to a new column (Column 4)
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # 4. Create a Bar Chart
    # We create a 'Reference' to tell the chart exactly which data to select
    values = Reference(sheet, 
                       min_row=2, 
                       max_row=sheet.max_row, 
                       min_col=4, 
                       max_col=4)
    
    chart = BarChart()
    chart.add_data(values)
    
    # Add the chart to the sheet, placing the top-left corner at cell 'e2'
    sheet.add_chart(chart, 'e2')

    # 5. Save the updated workbook
    # It's usually a good idea to save it as a new file during testing 
    # so you don't accidentally ruin your original data!
    wb.save(filename)
    print(f"Successfully processed and saved: {filename}")



# RUNNING THE AUTOMATION
# To test this, you need an actual Excel file (e.g., 'transactions.xlsx') 
# in the same folder as this Python script.

process_workbook('transactions.xlsx')